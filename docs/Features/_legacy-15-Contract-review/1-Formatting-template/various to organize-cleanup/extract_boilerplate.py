#!/usr/bin/env python3
"""
Valcre Workbook Boilerplate Extractor

Surgically extracts:
1. LISTS sheet text cells > 50 characters
2. Named ranges matching: Report_*, *Scope*, *Definition*, *Limiting*, *Certification*

Author: Python Pro
Date: 2025-12-11
"""

from pathlib import Path
from typing import Dict, List, Tuple
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def extract_lists_sheet_text(wb: Workbook, min_length: int = 50) -> List[Tuple[str, str]]:
    """
    Extract text cells from LISTS sheet with length > min_length.

    Args:
        wb: Opened workbook instance
        min_length: Minimum character length for text cells

    Returns:
        List of (cell_reference, text_value) tuples
    """
    results: List[Tuple[str, str]] = []

    if 'LISTS' not in wb.sheetnames:
        print("Warning: LISTS sheet not found in workbook")
        return results

    lists_sheet: Worksheet = wb['LISTS']

    # Iterate only over cells with values (data_only=True for formulas)
    for row in lists_sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and len(cell.value) > min_length:
                results.append((cell.coordinate, cell.value.strip()))

    return results


def extract_named_ranges(wb: Workbook) -> Dict[str, str]:
    """
    Extract named ranges matching specific patterns.

    Patterns: Report_*, *Scope*, *Definition*, *Limiting*, *Certification*

    Args:
        wb: Opened workbook instance

    Returns:
        Dict mapping range name to extracted text value
    """
    patterns = ['Report_', 'Scope', 'Definition', 'Limiting', 'Certification']
    results: Dict[str, str] = {}

    # Access workbook-level defined names
    for range_name, name_obj in wb.defined_names.items():
        # Check if name matches any pattern
        if any(
            range_name.startswith('Report_') or
            pattern.lower() in range_name.lower()
            for pattern in patterns
        ):
            try:
                # Get the destination (sheet, cell range)
                destinations = list(name_obj.destinations)
                if not destinations:
                    continue

                sheet_name, cell_range = destinations[0]

                # Handle both single cells and ranges
                if ':' in cell_range:
                    # Range - extract all cells
                    sheet = wb[sheet_name]
                    cells = sheet[cell_range]

                    # Flatten if nested (multi-row range)
                    values = []
                    if isinstance(cells[0], tuple):
                        for row in cells:
                            for cell in row:
                                if cell.value:
                                    values.append(str(cell.value))
                    else:
                        for cell in cells:
                            if cell.value:
                                values.append(str(cell.value))

                    results[range_name] = ' | '.join(values) if values else ''
                else:
                    # Single cell
                    sheet = wb[sheet_name]
                    cell = sheet[cell_range]
                    if cell.value:
                        results[range_name] = str(cell.value).strip()

            except Exception as e:
                print(f"Warning: Could not extract range '{range_name}': {e}")
                continue

    return results


def write_markdown_output(
    lists_data: List[Tuple[str, str]],
    named_ranges: Dict[str, str],
    output_path: Path
) -> None:
    """
    Write extracted data to markdown file.

    Args:
        lists_data: List of (cell_ref, text) tuples from LISTS sheet
        named_ranges: Dict of named range data
        output_path: Output markdown file path
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("# Valcre Workbook Boilerplate Extraction\n\n")
        f.write(f"**Source:** VAL251012 - North Battleford Apt, 1101, 1121 109 Street, North Battleford.xlsm\n")
        f.write(f"**Extraction Date:** 2025-12-11\n\n")

        # LISTS Sheet Section
        f.write("## LISTS Sheet - Text Cells (> 50 characters)\n\n")
        if lists_data:
            for cell_ref, text in lists_data:
                f.write(f"### {cell_ref}\n\n")
                f.write(f"{text}\n\n")
                f.write("---\n\n")
        else:
            f.write("*No text cells > 50 characters found in LISTS sheet*\n\n")

        # Named Ranges Section
        f.write("## Named Ranges\n\n")
        f.write("Patterns: `Report_*`, `*Scope*`, `*Definition*`, `*Limiting*`, `*Certification*`\n\n")

        if named_ranges:
            # Group by pattern
            report_ranges = {k: v for k, v in named_ranges.items() if k.startswith('Report_')}
            scope_ranges = {k: v for k, v in named_ranges.items() if 'scope' in k.lower()}
            definition_ranges = {k: v for k, v in named_ranges.items() if 'definition' in k.lower()}
            limiting_ranges = {k: v for k, v in named_ranges.items() if 'limiting' in k.lower()}
            cert_ranges = {k: v for k, v in named_ranges.items() if 'certification' in k.lower()}

            def write_range_section(title: str, ranges: Dict[str, str]) -> None:
                if ranges:
                    f.write(f"### {title}\n\n")
                    for name, value in sorted(ranges.items()):
                        f.write(f"**{name}:**\n\n")
                        f.write(f"{value}\n\n")
                        f.write("---\n\n")

            write_range_section("Report Ranges", report_ranges)
            write_range_section("Scope Ranges", scope_ranges)
            write_range_section("Definition Ranges", definition_ranges)
            write_range_section("Limiting Ranges", limiting_ranges)
            write_range_section("Certification Ranges", cert_ranges)
        else:
            f.write("*No matching named ranges found*\n\n")


def main() -> None:
    """Main extraction workflow."""
    # Paths
    workbook_path = Path("/Users/bencrowe/Development/APR-Dashboard-v3/docs/15-Contract-review/VAL251012 - North Battleford Apt, 1101, 1121 109 Street, North Battleford.xlsm")
    output_path = Path("/Users/bencrowe/Development/APR-Dashboard-v3/docs/15-Contract-review/BOILERPLATE-EXTRACTION.md")

    print(f"Opening workbook: {workbook_path.name}")
    print("Loading workbook (data_only=True to avoid formula execution)...")

    # Load workbook with data_only to get calculated values, not formulas
    wb = openpyxl.load_workbook(
        workbook_path,
        read_only=True,  # Memory efficient for large files
        data_only=True   # Get values, not formulas
    )

    print(f"Workbook loaded. Found {len(wb.sheetnames)} sheets.")

    # Extract LISTS sheet data
    print("\nExtracting LISTS sheet text cells (> 50 chars)...")
    lists_data = extract_lists_sheet_text(wb, min_length=50)
    print(f"Found {len(lists_data)} text cells in LISTS sheet")

    # Extract named ranges
    print("\nExtracting named ranges...")
    named_ranges = extract_named_ranges(wb)
    print(f"Found {len(named_ranges)} matching named ranges")

    # Write output
    print(f"\nWriting output to: {output_path}")
    write_markdown_output(lists_data, named_ranges, output_path)

    print("\n✅ Extraction complete!")
    print(f"   LISTS cells: {len(lists_data)}")
    print(f"   Named ranges: {len(named_ranges)}")
    print(f"   Output: {output_path}")


if __name__ == "__main__":
    main()
