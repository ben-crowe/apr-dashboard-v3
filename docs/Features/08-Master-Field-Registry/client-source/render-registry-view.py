#!/usr/bin/env python3
# Regenerates the in-app registry view from Chris's xlsx.
# Usage: python3 render-registry-view.py "<source.xlsx>"
# Excel-true row gutter: header = row 1, first data row = row 2 (matches openpyxl),
# so "Field Registry row 24 = ValueScenarios" lines up with the actual sheet.
import sys, html
from openpyxl import load_workbook
src = sys.argv[1]
wb = load_workbook(src, read_only=True, data_only=True)
tabs = wb.sheetnames

def cell(v):
    return "" if v is None else html.escape(str(v))

parts = []
parts.append("""<!doctype html><html><head><meta charset="utf-8">
<title>Valta Master Field Registry v03 — Chris's source (reviewed 2026-06-09)</title>
<style>
:root{--bd:#e3e6ea;--hd:#0f172a;--mut:#64748b;--accent:#2563eb;--hover:#f1f5f9;--gut:#eef2f6;}
*{box-sizing:border-box} body{font:13px/1.45 -apple-system,Segoe UI,Roboto,sans-serif;margin:0;color:#0f172a;background:#f8fafc}
header{padding:14px 18px;background:#fff;border-bottom:1px solid var(--bd)}
h1{font-size:15px;margin:0 0 2px} .sub{color:var(--mut);font-size:12px}
.src{margin-top:6px;font-size:11px;color:var(--mut)} .src code{background:#f1f5f9;padding:1px 5px;border-radius:4px}
.tabs{display:flex;gap:4px;flex-wrap:wrap;padding:10px 18px 0;background:#fff;border-bottom:1px solid var(--bd);position:sticky;top:0;z-index:8}
.tab{padding:6px 12px;border:1px solid var(--bd);border-bottom:none;border-radius:6px 6px 0 0;cursor:pointer;background:#f1f5f9;font-weight:500;color:#475569}
.tab.active{background:#fff;color:var(--accent);border-color:var(--accent);border-bottom:1px solid #fff;margin-bottom:-1px}
.panel{display:none;padding:14px 18px 40px} .panel.active{display:block}
table{border-collapse:collapse;width:100%;background:#fff;font-size:12px}
th,td{border:1px solid var(--bd);padding:4px 8px;text-align:left;vertical-align:top;max-width:320px;overflow-wrap:anywhere}
th{background:var(--hd);color:#fff;position:sticky;top:46px;font-weight:600;white-space:nowrap;z-index:4}
tr:nth-child(even) td{background:#fafbfc} tr:hover td{background:var(--hover)}
/* Excel-true row-number gutter — sticky on horizontal scroll */
td.rownum{position:sticky;left:0;background:var(--gut);color:var(--mut);font-weight:600;text-align:right;min-width:38px;white-space:nowrap;z-index:3}
tr:hover td.rownum{background:#e2e8f0}
th.rownum{position:sticky;left:0;top:46px;background:var(--hd);color:#fff;text-align:right;min-width:38px;z-index:6}
.count{color:var(--mut);font-weight:400;font-size:11px;margin-left:6px}
</style></head><body>
<header><h1>Valta Master Field Registry — v03 <span class="count">Chris's source · reviewed 2026-06-09</span></h1>
<div class="sub">Read-only mirror of the client's SharePoint workbook. Source of truth for which options a field shows and which fields map to Valcre. Left gutter = Excel-true row number (header = row 1).</div>
<div class="src">Source file: <code>docs/Features/08-Master-Field-Registry/client-source/Valta Master Field Registry v03 - reviewed 2026-06-09.xlsx</code></div></header>""")

parts.append('<div class="tabs">')
for i, t in enumerate(tabs):
    parts.append(f'<div class="tab{" active" if i==0 else ""}" onclick="show({i})">{html.escape(t)}</div>')
parts.append('</div>')

for i, t in enumerate(tabs):
    ws = wb[t]
    rows = [r for r in ws.iter_rows(values_only=True)]
    parts.append(f'<div class="panel{" active" if i==0 else ""}" id="p{i}">')
    parts.append('<table>')
    for ri, row in enumerate(rows):
        last = 0
        for ci, v in enumerate(row):
            if v is not None and str(v).strip() != "":
                last = ci
        cells = row[:last+1]
        if not any(("" if v is None else str(v).strip()) for v in cells) and ri != 0:
            # still number truly-blank rows so the gutter stays Excel-accurate
            cells = [""]
        excel_row = ri + 1  # openpyxl row 1 == header
        tag = "th" if ri == 0 else "td"
        gutter = f'<{tag} class="rownum">{excel_row}</{tag}>'
        body = "".join(f"<{tag}>{cell(v)}</{tag}>" for v in cells)
        parts.append(f"<tr>{gutter}{body}</tr>")
    parts.append('</table></div>')

parts.append("""<script>
function show(n){document.querySelectorAll('.tab').forEach((t,i)=>t.classList.toggle('active',i===n));
document.querySelectorAll('.panel').forEach((p,i)=>p.classList.toggle('active',i===n));}
</script></body></html>""")

out = "Valta-Registry-v03-VIEW.html"
open(out, "w").write("\n".join(parts))
print("WROTE", out, "·", len(tabs), "tabs · Excel-true row gutter added")
